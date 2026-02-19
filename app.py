#!/usr/bin/env python3
"""
銀行明細CSV → 月別Excelシート + 経営健康診断
Flask Webアプリ
"""

from flask import Flask, request, send_file, render_template_string, jsonify
import csv
import io
import zipfile
from datetime import datetime, date
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import json
import os
import tempfile

app = Flask(__name__)

# =====================================================
# 摘要パターン辞書（科目自動付与）
# 実際のCSVデータ（1,204件）を分析して作成
# =====================================================

# =====================================================
# 大カテゴリ定義 ── 会社正式科目体系（2026年2月版）
# =====================================================
CATEGORY_MAP = {
    # 収入
    "売上":           "🟢 売上",
    # 仕入
    "仕入":           "🔵 仕入",
    # 外注
    "外注費":         "🟡 外注費",
    # 人件費
    "人件費":         "🟠 人件費",
    # 販管費各種
    "交際費":         "🟣 交際費",
    "旅費交通費":     "🟣 旅費交通費",
    "車両費":         "🟣 車両費",
    "地代家賃":       "🟣 地代家賃",
    "光熱費":         "🟣 光熱費",
    "消耗品":         "🟣 消耗品",
    "通信費":         "🟣 通信費",
    "支払手数料":     "🟣 支払手数料",
    "福利厚生":       "🟣 福利厚生",
    "保険料":         "🟣 保険料",
    "租税公課":       "🟣 租税公課",
    "雑費":           "🟣 雑費",
    # 財務
    "事業借入":       "🔴 事業借入",
    "長期未払金":     "🔴 長期未払金",
    # 内部
    "口座振替":       "⚪ 振替",
    "現金引出":       "⚪ 振替",
    "受取利息":       "🟢 売上",
}

# =====================================================
# 仕訳辞書 ── 会社正式科目体系（全件精査済み）
# =====================================================

# ── 【内部振替・手数料】最優先 ──────────────────────────
OTHER_PATTERNS = {
    "振替 事業口座":            ("口座振替", "事業口座"),
    "振替 個人口座":            ("口座振替", "個人口座"),
    "カミデ　ケンタロウ":       ("口座振替", "資金移動"),  # 代表者口座間移動
    "ATM":                      ("現金引出", ""),
    "Mastercardデビット年会費": ("雑費",     "カード年会費"),
    "普通預金 利息":            ("受取利息", ""),
    "振込手数料":               ("支払手数料", "振込手数料"),
}

# ── 【仕入】 ──────────────────────────────────────────
PURCHASE_PATTERNS = {
    # 栄養補助・サプリメント
    "フリ−マンニユ−トラグル−プ":      ("仕入", "栄養補助食品"),
    "シゼンケンコウシヤ":              ("仕入", "健康食品"),
    "ＢＩＯ　ＡＣＴＩＶＥＳ　ＪＡＰＡＮ": ("仕入", "サプリ・輸入"),
    # 動物薬・医薬品
    "ニホンゼンヤクコウギヨウ":         ("仕入", "動物薬"),
    "ニホンゼンヤクコウギヨウカブシキ": ("仕入", "動物薬"),
    # 飼料・農業資材
    "エムピ−アグロ．カ":              ("仕入", "飼料・動物薬"),
    "MHF)MPｱｸﾞﾛ":                     ("仕入", "飼料・動物薬"),
    "MPｱｸﾞﾛ":                          ("仕入", "飼料・動物薬"),
    "カ）ムラカミキユウシヤ":          ("仕入", "飼料"),
    "ムラカミキユウシヤ":              ("仕入", "飼料"),
    # ITシステム（仕入原価扱い）
    "カ）シ−アイシ−フロンテイア":     ("仕入", "ITシステム"),
    # その他仕入
    "シグニ（カ":                      ("仕入", "仕入"),
    "シグニ":                          ("仕入", "仕入"),
    "ニホンガ−リツク":                ("仕入", "食材"),
}

# ── 【外注費】 ────────────────────────────────────────
OUTSOURCE_PATTERNS = {
    # 個人委託
    "カ）キタマ":                      ("外注費", "業務委託"),
    "サダモト　ユウイチ":              ("外注費", "業務委託"),
    "タグチ　カズオミ":                ("外注費", "業務委託"),
    "トミタ　ジユン":                  ("外注費", "業務委託"),
    "カ）マルサセンタ−":               ("外注費", "業務委託"),
    "ド）カリテイ−":                  ("外注費", "業務委託"),
    # 輸入・貿易
    "ナ−ヴイツク　インタ−ナシヨナル": ("外注費", "輸入・貿易"),
    "イ−ビ−エムトレ−デイング":       ("外注費", "輸入・貿易"),
    # 馬医療
    "カ）エクワインベツトグル−プ":     ("外注費", "馬医療"),
    # 専門家
    "ゼイ）スバルゴウドウカイケイ":    ("外注費", "税理士"),
    "ザイ）リユウツウシステムカイハツ": ("外注費", "システム開発"),
    "ザイ）セイブツカガクアンゼンケンキユウシヨ": ("外注費", "研究費"),
    # コンサル
    "クイ−ンビ−キヤピタル（カ":       ("外注費", "経営コンサル"),
    # その他外注
    "カ）アイレツクス":                ("外注費", "外注"),
    "カツヤマネクステ−ジ":             ("外注費", "外注"),
    "アナザ−レ−ン":                   ("外注費", "外注"),
}

# ── 【人件費（スタッフ名）】 ──────────────────────────
STAFF_NAMES = [
    "カミデ　サオリ",       # 家族スタッフ
    "オオムラ　トシノリ",   # スタッフ
    "コイズミ　ユカリ",     # スタッフ
    "ギヨウトク　マリア",   # スタッフ
    "キタハラ　ミユキ",     # スタッフ
    "サカモト　ミカ",       # スタッフ
    "タカハシ　ナオミ",     # スタッフ
    "サカイ　ケイコ",       # スタッフ
    "ハタケヤマ　ヨシカズ",  # スタッフ
    "ヤスカワ　エミコ",     # スタッフ
    "ナガクラ　ミスズ",     # スタッフ
    "ヤガミ　アヤノブ",     # スタッフ
    "サトウ　ユウナ",       # スタッフ
    "サトウ　カヨコ",       # スタッフ
    "ムトウ　アユミ",       # スタッフ
    "ヨシキ　カツヒコ",     # スタッフ
    "ナガタ　ユキヒロ",     # スタッフ
    "オオイシ　ユウヤ",     # スタッフ
    "ウエスギ　ユカリ",     # スタッフ
    "フジタ　カナ",         # スタッフ
    "タムラ　シンジ",       # スタッフ
    "サクライ　ハルミチ",   # スタッフ
    "ナガクラ　ミスズ",     # スタッフ
    "カンザキ　アユミ",     # スタッフ（Amex経由も含む）
    "オリベツト",           # スタッフ（カンザキ　アユミ）
]

# ── 【財務：事業借入・長期未払金】 ──────────────────────
FINANCE_PATTERNS = {
    # 事業借入（政策金融公庫）
    "ｶ)ﾆﾂﾎﾟﾝｾｲｻｸｷﾝﾕｳｺｳｺ":              ("事業借入", "政策金融公庫"),
    "カ）　ニツポンセイサクキンユウコウコ": ("事業借入", "政策金融公庫"),
    "ニツポンセイサクキンユウコウコ":       ("事業借入", "政策金融公庫"),
    # 長期未払金（クレジットカード・立替）
    "マネ−フオワ−ドケツサイ":             ("長期未払金", "クレジット"),
    "APｱﾌﾟﾗｽ":                            ("長期未払金", "クレジット"),
    "アメリカンエキスプレスインタ−ナシヨナルインコ−ポレイテツド": ("長期未払金", "Amex"),
    "アメリカンエキスプレスインタ−ナシヨナル": ("長期未払金", "Amex"),
    "ミツイスミトモカ−ド":                 ("長期未払金", "クレジット"),
}

# ── 【支払手数料】 ────────────────────────────────────
FEE_PATTERNS = {
    "カ）ベイフラワ−":    ("支払手数料", "販売手数料"),  # 花販売手数料
    "MHF)ﾔﾏﾄｳﾝﾕ":        ("支払手数料", "配送料"),
    "ﾔﾏﾄｳﾝﾕ":             ("支払手数料", "配送料"),
    "カ）シムネツト":      ("支払手数料", "通信・システム"),
}

# ── 【固定費各種】 ────────────────────────────────────
FIXED_PATTERNS = {
    "エアウオ−タ−ヒガシニホン": ("光熱費",   "ガス"),
    "フクシマトヨタ":            ("車両費",   ""),
    "ラクスル":                  ("消耗品",   "印刷・消耗品"),
}

# ── 【租税公課・その他】 ──────────────────────────────
MISC_PATTERNS = {
    "フクシマケンジユウイシカイ":           ("租税公課", "獣医師会"),
    "ゼンコクコウエイケイバジユウイシキヨウカイ": ("租税公課", "競馬獣医師会"),
    "ソウマキヨウタンカイ":                 ("租税公課", "業界団体"),
    "ジヤパンケネルクラブ":                 ("租税公課", "JKC"),
    "オダカシヨウコウカイ":                 ("租税公課", "商工会"),
    "ジ−ワンサラブレツドクラブ":           ("租税公課", "馬主クラブ"),
    "カワサキゴ−ゴ−ドツグクラブ":          ("租税公課", "犬クラブ"),
    "ヤマナシシンキン　エヌビ−シ−シユツパンキヨク": ("雑費", "出版"),
}

# ── 売上収入キーワード ─────────────────────────────────
SALES_KEYWORDS = [
    "チヤンピオンズフア−ム",
    "マウンテンビユ−ステ−ブル",
    "オイワケフア−ム",
    "ナストレ−ニングフア−ム",
    "アキバステ−ブル",
    "リクル−ト　ペイメント",
    "リクルート",
    "アマゾンジヤパン",
    "イ−ジ−　ラボ",
]

def guess_subject(description, amount_in=0, amount_out=0):
    """
    会社正式科目体系に基づく仕訳判定（方向優先ロジック）
    戻り値: (科目, 補助科目, 大カテゴリ, G列ラベル)
    """
    if not description:
        return ("", "", "", "")

    subject, sub = "", ""
    is_out = amount_out > 0
    is_in  = amount_in  > 0

    # ── STEP1: 方向問わず確定するもの（最優先）──
    for kw, result in OTHER_PATTERNS.items():
        if kw in description:
            subject, sub = result
            break

    # ── STEP2: 出金の科目判定 ──
    if not subject and is_out:

        # 外注費
        for kw, result in OUTSOURCE_PATTERNS.items():
            if kw in description:
                subject, sub = result
                break

        # 仕入
        if not subject:
            for kw, result in PURCHASE_PATTERNS.items():
                if kw in description:
                    subject, sub = result
                    break

        # 財務（事業借入・長期未払金）
        if not subject:
            for kw, result in FINANCE_PATTERNS.items():
                if kw in description:
                    subject, sub = result
                    break

        # 支払手数料
        if not subject:
            for kw, result in FEE_PATTERNS.items():
                if kw in description:
                    subject, sub = result
                    break

        # 固定費各種
        if not subject:
            for kw, result in FIXED_PATTERNS.items():
                if kw in description:
                    subject, sub = result
                    break

        # 租税公課・その他
        if not subject:
            for kw, result in MISC_PATTERNS.items():
                if kw in description:
                    subject, sub = result
                    break

        # 人件費（スタッフ名）
        if not subject:
            for name in STAFF_NAMES:
                if name in description:
                    subject, sub = "人件費", "スタッフ"
                    break

        # フォールバック
        if not subject:
            if "カ）" in description or "（カ" in description or "ゼイ）" in description or "ザイ）" in description:
                subject, sub = "外注費", "法人委託"
            else:
                subject, sub = "人件費", "スタッフ"

    # ── STEP3: 入金の科目判定 ──
    if not subject and is_in:
        subject, sub = "売上", ""
        # 補助科目で入金元を分類
        if any(kw in description for kw in ["ステ−ブル","フア−ム","チヤンピオンズ","マウンテン","アキバ","オイワケ","ナストレ−ニング"]):
            sub = "馬主・育成"
        elif any(kw in description for kw in ["リクル−ト","ペイメント","ＡＩＲペイ","デイ−ジ−フイナンシヤル"]):
            sub = "決済代行"
        elif "アマゾン" in description:
            sub = "Amazon"
        elif "ラボ" in description or "イ−ジ−　ラボ" in description:
            sub = "検査収入"
        elif "ベイフラワ" in description:
            sub = "花・装飾（売上）"
        elif "ＰＡＹＰＡＬ" in description:
            sub = "PayPal"

    # 大カテゴリ・G列ラベル生成
    category  = CATEGORY_MAP.get(subject, "⚪ その他")
    mid_label = sub if sub else subject
    g_label   = f"{category}  ›  {mid_label}" if subject else ""

    return (subject, sub, category, g_label)

# =====================================================
# CSV解析
# =====================================================
def parse_bank_csv(file_bytes):
    """銀行明細CSVを解析してデータリストを返す"""
    # エンコーディング自動検出
    for enc in ['shift_jis', 'cp932', 'utf-8-sig', 'utf-8']:
        try:
            text = file_bytes.decode(enc)
            break
        except:
            continue
    else:
        raise ValueError("CSVのエンコーディングを判定できませんでした")

    reader = csv.DictReader(io.StringIO(text))
    records = []
    
    # 列名のマッピング（表記ゆれ対応）
    col_map = {
        '日付': ['日付', 'date', '取引日'],
        '摘要': ['摘要', '内容', '取引内容', '摘　要'],
        '入金': ['入金金額', '入金', '入金額'],
        '出金': ['出金金額', '出金', '出金額'],
        '残高': ['残高'],
    }
    
    headers = reader.fieldnames or []
    
    def find_col(names):
        for n in names:
            if n in headers:
                return n
        return None
    
    col_date = find_col(col_map['日付'])
    col_desc = find_col(col_map['摘要'])
    col_in   = find_col(col_map['入金'])
    col_out  = find_col(col_map['出金'])
    col_bal  = find_col(col_map['残高'])
    
    if not col_date:
        raise ValueError("日付列が見つかりません")
    
    for row in reader:
        try:
            date_str = row.get(col_date, '').strip().replace('"', '')
            if not date_str:
                continue
            
            # 日付パース（YYYYMMDD or YYYY/MM/DD or YYYY-MM-DD）
            for fmt in ['%Y%m%d', '%Y/%m/%d', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    break
                except:
                    continue
            else:
                continue
            
            desc = row.get(col_desc, '').strip()
            
            def to_int(val):
                v = str(val or '').strip().replace(',', '').replace('"', '')
                return int(v) if v else 0
            
            amount_in  = to_int(row.get(col_in, 0))
            amount_out = to_int(row.get(col_out, 0))
            balance    = to_int(row.get(col_bal, 0))
            
            subject, sub_subject, category, g_label = guess_subject(desc, amount_in, amount_out)
            records.append({
                'date': dt,
                'year': dt.year,
                'month': dt.month,
                'day': dt.day,
                'description': desc,
                'amount_in': amount_in,
                'amount_out': amount_out,
                'balance': balance,
                'subject': subject,
                'sub_subject': sub_subject,
                'category': category,
                'g_label': g_label,
            })
        except Exception as e:
            continue
    
    return sorted(records, key=lambda x: x['date'])

# =====================================================
# Excel生成（既存GMO形式に準拠）
# =====================================================
MONTHS_JP = {
    1:'1月', 2:'2月', 3:'3月', 4:'4月',
    5:'5月', 6:'6月', 7:'7月', 8:'8月',
    9:'9月', 10:'10月', 11:'11月', 12:'12月'
}

def build_excel(records):
    """月別シートのExcelを生成"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシート削除
    
    # 月別にグループ化
    by_month = defaultdict(list)
    years = set()
    for r in records:
        by_month[(r['year'], r['month'])].append(r)
        years.add(r['year'])
    
    # スタイル定義
    header_font = Font(bold=True, name='Arial', size=11)
    data_font   = Font(name='Arial', size=10)
    title_fill  = PatternFill('solid', start_color='4472C4', end_color='4472C4')
    header_fill = PatternFill('solid', start_color='BDD7EE', end_color='BDD7EE')
    carry_fill  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
    total_fill  = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
    
    thin = Side(border_style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center')
    
    money_fmt = '#,##0'
    date_fmt  = '0'
    
    # 月の順番でシート作成
    sorted_months = sorted(by_month.keys())
    
    # 前月末残高を追跡
    prev_balance = {}
    
    for ym in sorted_months:
        year, month = ym
        month_records = by_month[ym]
        # D以降のアルファベットは sorted_months のインデックスで決定
        alpha = chr(ord("D") + sorted_months.index(ym))
        sheet_name = f"{alpha}. {year}年{MONTHS_JP[month]}"
        ws = wb.create_sheet(title=sheet_name)
        
        # ===== 行1: タイトル =====
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"E-MEITA仕訳Excel　{year}年{month}月"
        title_cell.font = Font(bold=True, name='Arial', size=13, color='FFFFFFFF')
        title_cell.fill = title_fill
        title_cell.alignment = center
        ws.row_dimensions[1].height = 22
        
        # ===== 行2: ヘッダー =====
        headers = ['日付', '摘　要', '出　　金', '入　　金', '残　　高', '科　目', '大分類  >  中分類（補助）']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, name='Arial', size=10, color='FF1F3864')
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[2].height = 18
        
        # ===== 行3: 前月繰越残高 =====
        prev_bal = prev_balance.get((year, month), None)
        if prev_bal is None:
            # 最初の月は最初のレコードの前残高を推定
            if month_records:
                first = month_records[0]
                if first['amount_in']:
                    prev_bal = first['balance'] - first['amount_in']
                elif first['amount_out']:
                    prev_bal = first['balance'] + first['amount_out']
                else:
                    prev_bal = 0
            else:
                prev_bal = 0
        
        ws['A3'] = None
        ws['B3'] = '前月繰越'
        ws['B3'].font = Font(bold=True, name='Arial', size=10)
        ws['B3'].fill = carry_fill
        ws['B3'].alignment = left
        ws['B3'].border = border
        
        ws['E3'] = prev_bal
        ws['E3'].number_format = money_fmt
        ws['E3'].font = Font(bold=True, name='Arial', size=10)
        ws['E3'].fill = carry_fill
        ws['E3'].alignment = right
        ws['E3'].border = border
        
        for c in [1, 3, 4, 6, 7]:
            cell = ws.cell(row=3, column=c)
            cell.fill = carry_fill
            cell.border = border
        
        ws.row_dimensions[3].height = 16
        
        # ===== 行4以降: 明細データ =====
        data_start_row = 4
        
        for i, rec in enumerate(month_records):
            row = data_start_row + i
            
            # A: 日付（日だけ）
            cell_a = ws.cell(row=row, column=1, value=rec['day'])
            cell_a.font = data_font
            cell_a.alignment = center
            cell_a.border = border
            cell_a.number_format = date_fmt
            
            # B: 摘要
            cell_b = ws.cell(row=row, column=2, value=rec['description'])
            cell_b.font = data_font
            cell_b.alignment = left
            cell_b.border = border
            
            # C: 出金
            cell_c = ws.cell(row=row, column=3, value=rec['amount_out'] if rec['amount_out'] else None)
            cell_c.font = data_font
            cell_c.alignment = right
            cell_c.border = border
            cell_c.number_format = money_fmt
            
            # D: 入金
            cell_d = ws.cell(row=row, column=4, value=rec['amount_in'] if rec['amount_in'] else None)
            cell_d.font = data_font
            cell_d.alignment = right
            cell_d.border = border
            cell_d.number_format = money_fmt
            
            # E: 残高（Excel数式）
            if i == 0:
                # 最初の行: 前月繰越 + 入金 - 出金
                formula = f"=E3+IF(D{row}>0,D{row},0)-IF(C{row}>0,C{row},0)"
            else:
                prev_row = row - 1
                formula = f"=E{prev_row}+IF(D{row}>0,D{row},0)-IF(C{row}>0,C{row},0)"
            
            cell_e = ws.cell(row=row, column=5, value=formula)
            cell_e.font = data_font
            cell_e.alignment = right
            cell_e.border = border
            cell_e.number_format = money_fmt
            
            # F: 科目（シンプル表記）
            cell_f = ws.cell(row=row, column=6, value=rec.get('subject', ''))
            cell_f.font = Font(name='Arial', size=9,
                               color='1F3864' if rec.get('subject') else 'BBBBBB')
            cell_f.alignment = left
            cell_f.border = border

            # G: 2階層カテゴリ「🟢 収入  >  売上収入（馬主・育成）」
            g_val = rec.get('g_label', '')
            cell_g = ws.cell(row=row, column=7, value=g_val)
            # 大分類ごとに文字色を変える
            # ARGB形式（FF+RGBの8桁）で指定しないと透明になる
            cat = rec.get('category', '')
            if   '収入' in cat: g_color = 'FF375623'
            elif '仕入' in cat: g_color = 'FF1F3864'
            elif '外注' in cat: g_color = 'FF7F6000'
            elif '人件' in cat: g_color = 'FFC55A11'
            elif '金融' in cat: g_color = 'FFC00000'
            elif '固定' in cat: g_color = 'FF4A148C'
            elif '振替' in cat: g_color = 'FF666666'
            else:               g_color = 'FF444444'
            cell_g.font      = Font(name='Arial', size=10, color=g_color)
            cell_g.alignment = left
            cell_g.border    = border
            
            ws.row_dimensions[row].height = 15
        
        # ===== 合計行 =====
        last_data_row = data_start_row + len(month_records) - 1
        total_row = last_data_row + 1
        
        if month_records:
            ws.merge_cells(f'A{total_row}:B{total_row}')
            cell_t = ws.cell(row=total_row, column=1, value='合　計')
            cell_t.font = Font(bold=True, name='Arial', size=10)
            cell_t.fill = total_fill
            cell_t.alignment = center
            cell_t.border = border
            
            # 出金合計
            ws.cell(row=total_row, column=3).value = f'=SUM(C{data_start_row}:C{last_data_row})'
            ws.cell(row=total_row, column=3).number_format = money_fmt
            ws.cell(row=total_row, column=3).font = Font(bold=True, name='Arial', size=10)
            ws.cell(row=total_row, column=3).fill = total_fill
            ws.cell(row=total_row, column=3).alignment = right
            ws.cell(row=total_row, column=3).border = border
            
            # 入金合計
            ws.cell(row=total_row, column=4).value = f'=SUM(D{data_start_row}:D{last_data_row})'
            ws.cell(row=total_row, column=4).number_format = money_fmt
            ws.cell(row=total_row, column=4).font = Font(bold=True, name='Arial', size=10)
            ws.cell(row=total_row, column=4).fill = total_fill
            ws.cell(row=total_row, column=4).alignment = right
            ws.cell(row=total_row, column=4).border = border
            
            # 月末残高
            ws.cell(row=total_row, column=5).value = f'=E{last_data_row}'
            ws.cell(row=total_row, column=5).number_format = money_fmt
            ws.cell(row=total_row, column=5).font = Font(bold=True, name='Arial', size=10, color='FFC00000')
            ws.cell(row=total_row, column=5).fill = total_fill
            ws.cell(row=total_row, column=5).alignment = right
            ws.cell(row=total_row, column=5).border = border
            
            for c in [6, 7]:
                ws.cell(row=total_row, column=c).fill = total_fill
                ws.cell(row=total_row, column=c).border = border
            
            ws.row_dimensions[total_row].height = 18
            
            # 翌月の前月残高を記録（月末残高）
            last_rec = month_records[-1]
            next_month = month + 1 if month < 12 else 1
            next_year  = year if month < 12 else year + 1
            prev_balance[(next_year, next_month)] = last_rec['balance']
        
        # ===== 列幅設定（A〜G全列が1画面に収まるよう設定）=====
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 22

        # ズーム80%に設定（全列が見えるよう）
        ws.sheet_view.zoomScale = 80

        # ヘッダー行を固定（スクロールしても1・2行目が常に見える）
        ws.freeze_panes = 'A3'

        # 印刷設定
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4
        ws.print_title_rows = '1:2'
    
    # ===== 集計・診断シートを先頭に挿入（index指定で順番固定）=====
    # 現在の月別シート数を取得してその前に挿入
    num_monthly = len(sorted_months)

    # 月別シートをいったん退避して後ろに移動
    # openpyxlはmove_sheetで順序変更できる
    build_summary_sheet(wb, records, by_month)   # 末尾に追加
    build_category_sheet(wb, records)            # 末尾に追加
    build_health_sheet(wb, records, by_month)    # 末尾に追加

    # シートを正しい順に並べ直す
    # 目標順: 📊年間サマリー, 📂カテゴリ別集計, 🏥経営健康診断, 月別(時系列)
    summary_sheets  = ['A. 📂カテゴリ別集計', 'B. 📊年間サマリー', 'C. 🏥経営健康診断']
    monthly_sheets  = [f"{chr(ord('D')+i)}. {y}年{MONTHS_JP[m]}" for i,(y,m) in enumerate(sorted_months)]
    desired_order   = summary_sheets + monthly_sheets

    for idx, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) * -1 + idx)

    # 最終的な順序を強制（move_sheetの累積ズレを修正）
    for idx, name in enumerate(desired_order):
        current_idx = wb.sheetnames.index(name)
        if current_idx != idx:
            wb.move_sheet(name, offset=idx - current_idx)

    return wb

def build_summary_sheet(wb, records, by_month):
    """年間サマリーシート"""
    ws = wb.create_sheet(title="B. 📊年間サマリー", index=0)
    
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right', vertical='center')
    
    header_fill = PatternFill('solid', start_color='1F3864', end_color='1F3864')
    alt_fill    = PatternFill('solid', start_color='F5F5F5', end_color='F5F5F5')
    total_fill  = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
    
    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = '年間入出金サマリー'
    ws['A1'].font = Font(bold=True, name='Arial', size=14, color='FFFFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1F3864', end_color='1F3864')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 24
    
    # ヘッダー
    headers = ['年月', '入金合計', '出金合計', '差引（純増減）', '月末残高', '入出金比率']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, name='Arial', size=10, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 18
    
    sorted_months = sorted(by_month.keys())
    money_fmt = '#,##0'
    pct_fmt   = '0.0%'
    
    total_in = total_out = 0
    
    for i, ym in enumerate(sorted_months):
        year, month = ym
        recs = by_month[ym]
        row = i + 3
        
        m_in  = sum(r['amount_in']  for r in recs)
        m_out = sum(r['amount_out'] for r in recs)
        m_bal = recs[-1]['balance'] if recs else 0
        m_diff = m_in - m_out
        m_ratio = m_in / m_out if m_out else None
        
        total_in  += m_in
        total_out += m_out
        
        fill = alt_fill if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
        
        ws.cell(row=row, column=1, value=f"{year}年{month}月").fill = fill
        ws.cell(row=row, column=2, value=m_in ).number_format = money_fmt
        ws.cell(row=row, column=3, value=m_out).number_format = money_fmt
        ws.cell(row=row, column=4, value=m_diff).number_format = money_fmt
        ws.cell(row=row, column=5, value=m_bal ).number_format = money_fmt
        ws.cell(row=row, column=6, value=m_ratio).number_format = pct_fmt if m_ratio else ''
        
        ratio_cell = ws.cell(row=row, column=6)
        if m_ratio:
            ratio_cell.value = m_ratio
            ratio_cell.number_format = '0.00'
            ratio_cell.value = round(m_ratio, 2)
        
        diff_cell = ws.cell(row=row, column=4)
        if m_diff < 0:
            diff_cell.font = Font(name='Arial', size=10, color='FFC00000')
        else:
            diff_cell.font = Font(name='Arial', size=10, color='FF375623')
        
        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            cell.alignment = right if c > 1 else center
            if c == 1:
                cell.alignment = center
    
    # 合計行
    total_row = len(sorted_months) + 3
    ws.cell(row=total_row, column=1, value='合　計').fill = total_fill
    ws.cell(row=total_row, column=2, value=total_in ).number_format = money_fmt
    ws.cell(row=total_row, column=3, value=total_out).number_format = money_fmt
    ws.cell(row=total_row, column=4, value=total_in - total_out).number_format = money_fmt
    ws.cell(row=total_row, column=5, value='')
    ws.cell(row=total_row, column=6, value='')
    
    for c in range(1, 7):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = total_fill
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.border = border
        cell.alignment = right if c > 1 else center
    
    # 列幅
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['F'].width = 12

def build_category_sheet(wb, records):
    """カテゴリ別集計シート"""
    from collections import defaultdict
    ws = wb.create_sheet(title="A. 📂カテゴリ別集計")

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    money_fmt = '#,##0'

    # カテゴリ別・科目別に集計
    cat_data   = defaultdict(lambda: {'in': 0, 'out': 0, 'count': 0})
    subj_data  = defaultdict(lambda: {'in': 0, 'out': 0, 'count': 0, 'category': ''})

    for r in records:
        cat  = r.get('category', '⚪ その他')
        subj = r.get('subject',  '未分類')
        cat_data[cat]['in']    += r['amount_in']
        cat_data[cat]['out']   += r['amount_out']
        cat_data[cat]['count'] += 1
        subj_data[subj]['in']       += r['amount_in']
        subj_data[subj]['out']      += r['amount_out']
        subj_data[subj]['count']    += 1
        subj_data[subj]['category'] = cat

    # ===== タイトル =====
    ws.merge_cells('A1:F1')
    ws['A1'] = '📂 カテゴリ別 入出金集計'
    ws['A1'].font  = Font(bold=True, name='Arial', size=13, color='FFFFFFFF')
    ws['A1'].fill  = PatternFill('solid', start_color='203864', end_color='203864')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 24

    # カテゴリの並び順
    CAT_ORDER = [
        "🟢 収入",
        "🔵 仕入・材料費",
        "🟡 外注・業務委託",
        "🟠 人件費",
        "🔴 金融・借入",
        "🟣 固定費",
        "⚪ 振替・内部",
        "⚪ その他",
    ]
    # 科目 → カテゴリのリバース引き引き
    subj_to_cat = {s: d['category'] for s, d in subj_data.items()}

    # カテゴリ色定義
    CAT_COLORS = {
        "🟢 収入":         ('E2EFDA', '375623'),
        "🔵 仕入・材料費":  ('DEEAF1', '1F3864'),
        "🟡 外注・業務委託": ('FFFACD', '7F6000'),
        "🟠 人件費":        ('FCE4D6', 'C55A11'),
        "🔴 金融・借入":    ('FFE0E0', 'C00000'),
        "🟣 固定費":        ('EDE7F6', '4A148C'),
        "⚪ 振替・内部":    ('F5F5F5', '555555'),
        "⚪ その他":        ('F5F5F5', '555555'),
    }

    row = 2

    # ===== 大カテゴリ別サマリー表 =====
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 大カテゴリ別サマリー'
    ws[f'A{row}'].font = Font(bold=True, name='Arial', size=11, color='FF1F3864')
    ws[f'A{row}'].fill = PatternFill('solid', start_color='DEEAF1', end_color='DEEAF1')
    ws.row_dimensions[row].height = 20
    row += 1

    hdr = ['カテゴリ', '入金合計', '出金合計', '差引', '件数', '']
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font  = Font(bold=True, name='Arial', size=10, color='FFFFFFFF')
        cell.fill  = PatternFill('solid', start_color='1F3864', end_color='1F3864')
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 18
    row += 1

    total_in = total_out = 0
    for cat in CAT_ORDER:
        d = cat_data.get(cat)
        if not d:
            continue
        bg, fg = CAT_COLORS.get(cat, ('F5F5F5', '333333'))
        fill = PatternFill('solid', start_color=bg, end_color=bg)
        diff = d['in'] - d['out']

        vals = [cat, d['in'], d['out'], diff, d['count'], '']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill   = fill
            cell.border = border
            cell.alignment = right if c > 1 else left
            if c in (2,3,4):
                cell.number_format = money_fmt
                cell.font = Font(bold=True, name='Arial', size=10, color=fg)
            elif c == 1:
                cell.font = Font(bold=True, name='Arial', size=10, color=fg)
            else:
                cell.font = Font(name='Arial', size=10, color=fg)
            if c == 4 and diff < 0:
                cell.font = Font(bold=True, name='Arial', size=10, color='FFC00000')
        total_in  += d['in']
        total_out += d['out']
        ws.row_dimensions[row].height = 17
        row += 1

    # 合計行
    total_fill = PatternFill('solid', start_color='203864', end_color='203864')
    for c, v in enumerate(['合　計', total_in, total_out, total_in-total_out, '', ''], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill   = total_fill
        cell.font   = Font(bold=True, name='Arial', size=11, color='FFFFFFFF')
        cell.border = border
        cell.alignment = right if c > 1 else center
        if c in (2,3,4):
            cell.number_format = money_fmt
    ws.row_dimensions[row].height = 20
    row += 2

    # ===== 科目別明細表 =====
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 科目別明細'
    ws[f'A{row}'].font = Font(bold=True, name='Arial', size=11, color='FF1F3864')
    ws[f'A{row}'].fill = PatternFill('solid', start_color='DEEAF1', end_color='DEEAF1')
    ws.row_dimensions[row].height = 20
    row += 1

    hdr2 = ['カテゴリ', '科目', '入金合計', '出金合計', '差引', '件数']
    for c, h in enumerate(hdr2, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font  = Font(bold=True, name='Arial', size=10, color='FFFFFFFF')
        cell.fill  = PatternFill('solid', start_color='1F3864', end_color='1F3864')
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 18
    row += 1

    # カテゴリ順に科目をソート
    sorted_subjs = []
    for cat in CAT_ORDER:
        for subj, d in sorted(subj_data.items(), key=lambda x: -x[1]['out']-x[1]['in']):
            if d['category'] == cat:
                sorted_subjs.append((cat, subj, d))

    prev_cat = None
    for i, (cat, subj, d) in enumerate(sorted_subjs):
        bg, fg = CAT_COLORS.get(cat, ('F5F5F5', '333333'))
        if cat != prev_cat:
            fill = PatternFill('solid', start_color=bg, end_color=bg)
        else:
            # 同カテゴリは少し薄く
            fill = PatternFill('solid', start_color='FAFAFA' if i%2==0 else 'F0F0F0', end_color='FAFAFA' if i%2==0 else 'F0F0F0')
        prev_cat = cat

        diff = d['in'] - d['out']
        # 科目に補助科目を付加して表示
        subj_sample = next((r.get('sub_subject','') for r in records if r['subject']==subj and r.get('sub_subject','')), '')
        subj_display2 = f"{subj}（{subj_sample}）" if subj_sample else subj
        vals = [cat, subj_display2, d['in'], d['out'], diff, d['count']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill   = fill
            cell.border = border
            cell.alignment = right if c > 2 else left
            if c in (3,4,5):
                cell.number_format = money_fmt
                cell.font = Font(name='Arial', size=10)
            else:
                cell.font = Font(name='Arial', size=10)
            if c == 5 and diff < 0:
                cell.font = Font(name='Arial', size=10, color='FFC00000')
        ws.row_dimensions[row].height = 16
        row += 1

    # 列幅
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 8

def build_health_sheet(wb, records, by_month):
    """経営健康診断シート（動物病院モード × BizClinic参照ベンチマーク）"""
    from collections import defaultdict
    ws = wb.create_sheet(title="C. 🏥経営健康診断")

    thin   = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')
    wrap   = Alignment(horizontal='left',   vertical='top', wrap_text=True)

    # ===== 動物病院ベンチマーク（政府統計・BizClinic準拠） =====
    BM = {
        '仕入率':      {'q1': 0.22, 'median': 0.27, 'q3': 0.33, 'lower_better': True},
        '人件費率':    {'q1': 0.28, 'median': 0.34, 'q3': 0.40, 'lower_better': True},
        '固定費率':    {'q1': 0.18, 'median': 0.22, 'q3': 0.27, 'lower_better': True},
        '営業利益率':  {'q1': 0.05, 'median': 0.12, 'q3': 0.20, 'lower_better': False},
    }

    def bm_rank(value, bm_key):
        bm = BM[bm_key]
        q1, med, q3, lb = bm['q1'], bm['median'], bm['q3'], bm['lower_better']
        if lb:
            if value <= q1:   return 'top',         '🏆 上位25%（優秀）',     'FF1B4F2A'
            if value <= med:  return 'above_median', '✅ 中央値以上（良好）',   'FF375623'
            if value <= q3:   return 'below_median', '⚠️ 中央値以下（要注意）','FF7F6000'
            return            'bottom',              '🔴 下位25%（要改善）',   'FFC00000'
        else:
            if value >= q3:   return 'top',         '🏆 上位25%（優秀）',     'FF1B4F2A'
            if value >= med:  return 'above_median', '✅ 中央値以上（良好）',   'FF375623'
            if value >= q1:   return 'below_median', '⚠️ 中央値以下（要注意）','FF7F6000'
            return            'bottom',              '🔴 下位25%（要改善）',   'FFC00000'

    def deviation_pct(value, median):
        return (value - median) / median * 100 if median else 0

    # ===== KPI集計 =====
    total_in  = sum(r['amount_in']  for r in records)
    total_out = sum(r['amount_out'] for r in records)
    net       = total_in - total_out
    months_count = len(by_month)

    by_subj = defaultdict(int)
    top_items_by_subj = defaultdict(list)  # 科目→[(金額,説明)]
    for r in records:
        if r['amount_out']:
            s = r.get('subject','')
            amt = r['amount_out']
            by_subj[s] += amt
            top_items_by_subj[s].append((amt, r['description'][:20]))

    sales    = total_in
    cogs     = by_subj.get('仕入', 0)
    labor    = by_subj.get('人件費', 0)
    outsrc   = by_subj.get('外注費', 0)
    fixed    = (by_subj.get('光熱費', 0) + by_subj.get('通信費', 0) +
                by_subj.get('車両費', 0) + by_subj.get('地代家賃', 0) +
                by_subj.get('消耗品', 0) + by_subj.get('保険料', 0) +
                by_subj.get('雑費', 0)   + by_subj.get('租税公課', 0) +
                by_subj.get('福利厚生', 0))
    labor_total = labor + outsrc  # 人件費＋外注費
    fee      = by_subj.get('支払手数料', 0)
    borrow   = by_subj.get('事業借入', 0)
    transfer = by_subj.get('口座振替', 0)

    cogs_r    = cogs      / sales if sales else 0
    labor_r   = labor_total / sales if sales else 0
    fixed_r   = fixed     / sales if sales else 0
    fee_r     = fee       / sales if sales else 0
    fl_ratio  = cogs_r + labor_r
    op_margin = 1 - cogs_r - labor_r - fixed_r - fee_r

    # 月別収支
    monthly_data = []
    for ym in sorted(by_month.keys()):
        recs = by_month[ym]
        m_in  = sum(r['amount_in']  for r in recs)
        m_out = sum(r['amount_out'] for r in recs)
        monthly_data.append({'ym': ym, 'in': m_in, 'out': m_out, 'diff': m_in - m_out})
    red_months = sum(1 for m in monthly_data if m['diff'] < 0)
    avg_in  = total_in  / months_count if months_count else 0
    avg_out = total_out / months_count if months_count else 0

    # 科目ごとのTop取引先
    def top_vendors(subj_key, n=3):
        items = sorted(top_items_by_subj.get(subj_key, []), reverse=True)[:n]
        return items

    # 主因分析（BizClinic cause_analysis参照）
    scores = {
        '仕入率':   max(0, cogs_r   - BM['仕入率']['q3']),
        '人件費率': max(0, labor_r  - BM['人件費率']['q3']),
        '固定費率': max(0, fixed_r  - BM['固定費率']['q3']),
    }
    primary_cause = max(scores, key=lambda k: scores[k]) if any(v>0 for v in scores.values()) else None

    # ===== スタイル定義 =====
    def fill(hex6):
        c = hex6 if len(hex6) == 8 else 'FF' + hex6
        return PatternFill('solid', start_color=c, end_color=c)
    DARK_GREEN = '1B4F2A';  LIGHT_GREEN = 'E8F5E9'
    YELLOW_BG  = 'FFFFF9C4'; RED_BG     = 'FFEBEE'
    GRAY_BG    = 'F8F8F8';   WHITE      = 'FFFFFF'

    def write_section(row, title, bg=DARK_GREEN, fg='FFFFFFFF'):
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, name='Arial', size=11, color=fg)
        ws[f'A{row}'].fill = fill(bg)
        ws[f'A{row}'].alignment = left
        ws.row_dimensions[row].height = 22

    def write_table_header(row, headers, col_widths=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font  = Font(bold=True, name='Arial', size=9, color='FFFFFFFF')
            cell.fill  = fill(DARK_GREEN)
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[row].height = 16

    def write_kpi_row(row, label, actual, actual_str, median, q1, q3,
                      rank_key, rank_label, rank_color, advice, even=True):
        bg = GRAY_BG if even else WHITE
        dev = deviation_pct(actual, median)

        cells_data = [
            (1, label,       left,   bg, Font(name='Arial', size=10)),
            (2, actual_str,  right,  bg, Font(bold=True, name='Arial', size=10)),
            (3, f'{median:.1%}',  center, bg, Font(name='Arial', size=9, color='FF888888')),
            (4, f'{q1:.1%}',     center, bg, Font(name='Arial', size=9, color='FF4CAF50')),
            (5, f'{q3:.1%}',     center, bg, Font(name='Arial', size=9, color='FFEF5350')),
            (6, f'{dev:+.1f}%',  center, bg, Font(name='Arial', size=9,
                                    color='FF4CAF50' if dev < 0 and actual < median else
                                           'FFEF5350' if dev > 5 else 'FF333333')),
            (7, rank_label,  center, bg, Font(bold=True, name='Arial', size=9, color=rank_color)),
        ]
        for c, val, aln, bg_col, fnt in cells_data:
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = fnt;  cell.fill = fill(bg_col)
            cell.alignment = aln; cell.border = border

        # アドバイス（結合セル）
        ws.merge_cells(f'H{row}:I{row}')
        adv_cell = ws.cell(row=row, column=8, value=advice)
        adv_bg = 'E8F5E9' if 'top' in rank_key or 'above' in rank_key else 'FFF9C4' if 'below' in rank_key else 'FFEBEE'
        adv_cell.font = Font(name='Arial', size=9,
                             color='FF1B4F2A' if 'top' in rank_key or 'above' in rank_key else
                                    'FF7F6000' if 'below' in rank_key else 'FFC00000')
        adv_cell.fill = fill(adv_bg)
        adv_cell.alignment = wrap
        adv_cell.border = border
        ws.cell(row=row, column=9).border = border
        ws.row_dimensions[row].height = 36

    # ===== ヘッダー =====
    ws.merge_cells('A1:I1')
    ws['A1'] = 'E-MEITA仕訳Excel　🏥 経営健康診断レポート（動物病院モード）'
    ws['A1'].font  = Font(bold=True, name='Arial', size=14, color='FFFFFFFF')
    ws['A1'].fill  = fill(DARK_GREEN)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:I2')
    period = f"{records[0]['date'].strftime('%Y/%m/%d')} ～ {records[-1]['date'].strftime('%Y/%m/%d')}" if records else 'N/A'
    ws['A2'] = f'生成: {datetime.now().strftime("%Y/%m/%d %H:%M")}  |  対象: {period}  |  {months_count}ヶ月  |  ベンチマーク: 政府統計（財務省・中小企業庁 2023年度）'
    ws['A2'].font = Font(name='Arial', size=8, color='FF888888')
    ws['A2'].alignment = left

    # ===== セクション1: 動物病院KPI比較 =====
    row = 4
    write_section(row, '■ 動物病院 経営KPI  ─  業界ベンチマーク比較（政府統計）')
    row += 1

    write_table_header(row, ['KPI', '実績値', '中央値', 'Q1(優)', 'Q3(警)', '中央値比', '判定', '備考・アドバイス（2〜3行）'])
    row += 1

    # 仕入率
    rank, rl, rc = bm_rank(cogs_r, '仕入率')
    top_c = top_vendors('仕入', 3)
    top_txt = ' / '.join([f'{d[:12]}({a/10000:.0f}万)' for a,d in top_c]) if top_c else ''
    if rank in ('top', 'above_median'):
        adv = f'✅ 仕入コストは適正水準。\n主要: {top_txt}\n引き続き仕入先との価格交渉を継続。'
    elif rank == 'below_median':
        adv = f'⚠️ 仕入率が中央値を超えています。\n主要: {top_txt}\n単価上昇か使用量増加かを切り分けてください。フリーマンニュートラ等の契約内容を確認。'
    else:
        adv = f'🔴 仕入率が業界下位25%です。緊急見直しが必要。\n主要: {top_txt}\n複数社見積比較・在庫最適化・廃棄削減を同時に実施してください。'
    write_kpi_row(row, '仕入率（仕入/売上）', cogs_r, f'{cogs_r:.1%}  (¥{cogs:,})',
                  0.27, 0.22, 0.33, rank, rl, rc, adv, True)
    row += 1

    # 人件費率（人件費＋外注費）
    rank, rl, rc = bm_rank(labor_r, '人件費率')
    top_l = top_vendors('人件費', 3)
    top_ltxt = ' / '.join([f'{d[:10]}({a/10000:.0f}万)' for a,d in top_l]) if top_l else ''
    if rank in ('top', 'above_median'):
        adv = f'✅ 人件費は適正水準。\n主要スタッフ: {top_ltxt}\n現状の人員配置と売上規模のバランスは良好です。'
    elif rank == 'below_median':
        adv = f'⚠️ 人件費率が中央値を超えています。\n主要: {top_ltxt}\n診療件数・売上と各スタッフ稼働の紐付けを確認してください。非繁忙時間帯のシフト最適化を検討。'
    else:
        adv = f'🔴 人件費率が業界下位25%です。\n主要: {top_ltxt}\n売上1万円あたりの人件費を計算し、診療単価引上げ・時間当たり生産性の見直しを最優先に。'
    write_kpi_row(row, '人件費率（人件費＋外注/売上）', labor_r, f'{labor_r:.1%}  (¥{labor_total:,})',
                  0.34, 0.28, 0.40, rank, rl, rc, adv, False)
    row += 1

    # 固定費率
    rank, rl, rc = bm_rank(fixed_r, '固定費率')
    if rank in ('top', 'above_median'):
        adv = f'✅ 固定費は適正水準（¥{fixed:,}）。\n現状の固定費管理は良好。年1回は各契約の見直しをルーティン化しましょう。'
    elif rank == 'below_median':
        adv = f'⚠️ 固定費率がやや高め（¥{fixed:,}）。\n光熱費・通信費・会費のうち最大項目から見直し。一度に全部変えず、1項目ずつ確認してください。'
    else:
        adv = f'🔴 固定費率が高水準（¥{fixed:,}）。\n各契約の単価・頻度・必要性を精査。すぐに解約できない契約でも、条件変更交渉は可能な場合があります。'
    write_kpi_row(row, '固定費率（光熱・通信・会費等/売上）', fixed_r, f'{fixed_r:.1%}  (¥{fixed:,})',
                  0.22, 0.18, 0.27, rank, rl, rc, adv, True)
    row += 1

    # 営業利益率
    rank, rl, rc = bm_rank(op_margin, '営業利益率')
    if op_margin >= 0.15:
        adv = f'🏆 営業利益率{op_margin:.1%}は業界上位水準です。\n現在の価格設定・コスト管理の仕組みを維持・標準化してください。利益を設備投資や人材育成へ再投資する時期です。'
    elif op_margin >= 0.12:
        adv = f'✅ 営業利益率{op_margin:.1%}は中央値水準です。\n主因を改善すれば上位25%（20%超）を目指せます。'
    elif op_margin >= 0.05:
        adv = f'⚠️ 営業利益率{op_margin:.1%}は業界平均以下です。\n診療単価の見直しまたは集患強化で売上を伸ばしながら固定費比率を下げる戦略を検討。'
    else:
        adv = f'🔴 営業利益率{op_margin:.1%}は危険水準です。\n損益分岐点を計算し、最低必要売上額を把握した上で価格設定を見直してください。'
    write_kpi_row(row, '営業利益率（推定）', op_margin, f'{op_margin:.1%}',
                  0.12, 0.05, 0.20, rank, rl, rc, adv, False)
    row += 2

    # ===== セクション2: FL比率（BizClinic参照） =====
    write_section(row, '■ FL比率分析  ─  仕入率＋人件費率（食材・人の合計負担）', bg='2E4057')
    row += 1

    fl_bg = LIGHT_GREEN if fl_ratio <= 0.61 else YELLOW_BG if fl_ratio <= 0.70 else 'FFEBEE'
    fl_color = 'FF1B4F2A' if fl_ratio <= 0.61 else 'FF7F6000' if fl_ratio <= 0.70 else 'FFC00000'
    fl_status = '✅ 健全' if fl_ratio <= 0.61 else '⚠️ やや高め' if fl_ratio <= 0.70 else '🔴 危険水準'

    ws.merge_cells(f'A{row}:I{row}')
    heavier = '仕入コスト' if cogs_r > labor_r else '人件費'
    fl_msg = (f'{fl_status}  FL比率: {fl_ratio:.1%}  （仕入率{cogs_r:.1%} ＋ 人件費率{labor_r:.1%}）   '
              f'{"▶ 適正範囲（61%以内）" if fl_ratio <= 0.61 else f"▶ {heavier}が重い。まずこちらから改善してください。"}')
    ws[f'A{row}'] = fl_msg
    ws[f'A{row}'].font = Font(bold=True, name='Arial', size=11, color=fl_color)
    ws[f'A{row}'].fill = fill(fl_bg)
    ws[f'A{row}'].alignment = left
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 22
    row += 2

    # ===== セクション3: 月別推移 =====
    write_section(row, '■ 月別入出金推移')
    row += 1
    write_table_header(row, ['年月', '入金', '出金', '差引', '判定', '入金比較（前月差）', '', '', ''])
    row += 1

    prev_in = None
    for i, md in enumerate(monthly_data):
        ym = md['ym']
        diff_in = md['in'] - prev_in if prev_in else 0
        diff_txt = f'{diff_in:+,}' if prev_in else '─'
        ratio = md['in'] / md['out'] if md['out'] > 0 else 9
        if md['diff'] > 0 and ratio > 1.20: judge, jc = '✅ 黒字', 'FF375623'
        elif md['diff'] > 0:                judge, jc = '⚠️ 微益', 'FF7F6000'
        else:                               judge, jc = '🔴 赤字', 'FFC00000'

        bg_row = GRAY_BG if i%2==0 else WHITE
        for c, v, aln, fmt in [
            (1, f"{ym[0]}年{ym[1]}月", center, None),
            (2, md['in'],  right, '#,##0'),
            (3, md['out'], right, '#,##0'),
            (4, md['diff'], right, '#,##0'),
            (5, judge, center, None),
            (6, diff_txt, center, None),
        ]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(bg_row); cell.border = border; cell.alignment = aln
            if fmt: cell.number_format = fmt
            if c == 4: cell.font = Font(name='Arial', size=10, color='FFC00000' if md['diff']<0 else 'FF333333')
            if c == 5: cell.font = Font(bold=True, name='Arial', size=9, color=jc)
        for c in [7, 8, 9]:
            ws.cell(row=row, column=c).fill = fill(bg_row); ws.cell(row=row, column=c).border = border
        ws.row_dimensions[row].height = 16
        prev_in = md['in']
        row += 1

    row += 1

    # ===== セクション4: 総合スコア =====
    write_section(row, '■ 総合経営健康スコア  ─  BizClinic準拠アルゴリズム')
    row += 1

    score = 100
    all_advice = []

    # 主因優先のアドバイス（report.py _build_actions参照）
    kpi_checks = [
        ('仕入率',   cogs_r,   0.27, 0.33, '仕入・原価管理'),
        ('人件費率', labor_r,  0.34, 0.40, '人件費・生産性'),
        ('固定費率', fixed_r,  0.22, 0.27, '固定費削減'),
        ('営業利益率', op_margin, 0.12, 0.05, '収益改善'),
    ]
    for kpi_name, val, median, warn_th, theme in kpi_checks:
        lb = kpi_name != '営業利益率'
        over = val > warn_th if lb else val < warn_th
        critical = (val > warn_th * 1.2) if lb else (val < warn_th * 0.5)
        if critical:
            score -= 20
            all_advice.append(('🔴', f'{kpi_name}が危険水準です（{val:.1%} / 目安{warn_th:.1%}）。{theme}を最優先で改善してください。'))
        elif over:
            score -= 8
            all_advice.append(('⚠️', f'{kpi_name}がやや高め（{val:.1%} / 中央値{median:.1%}）。{theme}の見直しを検討してください。'))

    if red_months > 0:
        score -= min(15, red_months * 5)
        all_advice.append(('⚠️' if red_months <= 2 else '🔴',
                           f'赤字月が{red_months}ヶ月あります。月次収支管理の強化と、赤字月の支出パターンを確認してください。'))

    if fl_ratio > 0.70:
        score -= 10
        all_advice.append(('🔴', f'FL比率{fl_ratio:.1%}が70%超。{"仕入コスト" if cogs_r > labor_r else "人件費"}の削減を先行させてください。'))

    if net > 0 and len(all_advice) == 0:
        all_advice.append(('✅', '全KPIが適正範囲内です。現在の経営スタイルと価格設定を維持してください。利益を設備・人材育成へ再投資する段階です。'))

    score = max(0, min(100, score))
    if score >= 80:   se, sl, sc_col = '🏆', '優秀', 'FF1B4F2A'
    elif score >= 60: se, sl, sc_col = '⚠️', '要注意', 'FF7F6000'
    else:             se, sl, sc_col = '🚨', '要改善', 'FFC00000'

    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = f'{se}  総合スコア：{score}点 / 100点  ({sl})   主因: {primary_cause or "良好・問題なし"}'
    ws[f'A{row}'].font = Font(bold=True, name='Arial', size=14, color=sc_col)
    ws[f'A{row}'].fill = fill('E8F5E9' if score >= 80 else 'FFF9C4' if score >= 60 else 'FFEBEE')
    ws[f'A{row}'].alignment = center
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=c).fill = fill('E8F5E9' if score >= 80 else 'FFF9C4' if score >= 60 else 'FFEBEE')
    ws.row_dimensions[row].height = 28
    row += 2

    for emoji_a, text_a in all_advice:
        bg_a = 'E8F5E9' if emoji_a == '✅' else 'FFF9C4' if emoji_a == '⚠️' else 'FFEBEE'
        tc_a = 'FF1B4F2A' if emoji_a == '✅' else 'FF7F6000' if emoji_a == '⚠️' else 'FFC00000'
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = f'  {emoji_a}  {text_a}'
        ws[f'A{row}'].font = Font(name='Arial', size=10, color=tc_a)
        ws[f'A{row}'].fill = fill(bg_a)
        ws[f'A{row}'].alignment = wrap
        ws[f'A{row}'].border = border
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = fill(bg_a)
            ws.cell(row=row, column=c).border = border
        ws.row_dimensions[row].height = 28
        row += 1

    # ===== 列幅・設定 =====
    widths = {'A':22,'B':14,'C':10,'D':9,'E':9,'F':9,'G':10,'H':28,'I':5}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = 'A3'

# =====================================================
# Flask ルーティング
# =====================================================

# =====================================================
# 月次P&L評価エンジン
# =====================================================

# 業界ベンチマーク（動物病院・馬産業）
BENCHMARKS = {
    '仕入比率':    {'median': 0.28, 'good': 0.25, 'warn': 0.38, 'label': '仕入・原材料費'},
    '人件費比率':  {'median': 0.28, 'good': 0.25, 'warn': 0.40, 'label': '人件費（外注含む）'},
    '固定費比率':  {'median': 0.12, 'good': 0.10, 'warn': 0.20, 'label': '固定費（家賃・光熱・通信）'},
    '営業利益率':  {'median': 0.18, 'good': 0.25, 'warn': 0.08, 'label': '営業利益率', 'higher_is_better': True},
    '粗利率':     {'median': 0.65, 'good': 0.70, 'warn': 0.50, 'label': '粗利率',     'higher_is_better': True},
}

# 科目 → P&Lカテゴリのマッピング
PL_CATEGORY = {
    '仕入':      'cogs',
    '外注費':    'labor',
    '人件費':    'labor',
    '福利厚生':  'labor',
    '地代家賃':  'fixed',
    '光熱費':    'fixed',
    '通信費':    'fixed',
    '車両費':    'fixed',
    '交際費':    'selling',
    '旅費交通費': 'selling',
    '消耗品':    'other',
    '支払手数料': 'other',
    '租税公課':  'tax',
    '雑費':      'other',
    '保険料':    'fixed',
    '事業借入':  'financing',
    '長期未払金': 'financing',
}

import re as _re

def parse_pl_text(text: str) -> dict:
    """月次P&Lテキストを解析して数値を抽出する"""
    result = {
        'period': '',
        'revenue': 0,
        'items': {},      # 科目名: 金額
        'raw_text': text,
    }
    
    # 期間抽出
    m = _re.search(r'(20\d{2})[年/\-](\d{1,2})月?', text)
    if m:
        result['period'] = f"{m.group(1)}年{int(m.group(2))}月"
    
    # テキスト正規化
    t = text.replace('　', ' ').replace('：', ':').replace(',', '').replace('¥', '').replace('円', '').replace('￥', '')
    
    # 売上を抽出
    for kw in ['売上', '売上高', '収入', '入金']:
        pm = _re.search(rf'{kw}[\s:：　]*([\d,]+)', t)
        if pm:
            try:
                result['revenue'] = int(pm.group(1).replace(',',''))
                break
            except:
                pass
    
    # 科目と金額を抽出（「科目名　金額」形式）
    seen_items = {}
    for line in t.split('\n'):
        line = line.strip()
        if not line:
            continue
        # パターン: 「科目名 数字」
        pm = _re.match(r'^([^\d\n:]{1,15})\s+(\d+)\s*$', line)
        if pm:
            key = pm.group(1).strip()
            skip_keys = {'入金', '出金', '合計', '収入', '売上', '売上高', '出費配分内訳', '入出費配分内訳'}
            if key not in skip_keys and len(key) >= 2:
                val = int(pm.group(2))
                # 重複は最初を採用
                if key not in seen_items:
                    seen_items[key] = val
    
    result['items'] = seen_items
    return result


def evaluate_pl(pl_data: dict, bank_records: list = None) -> dict:
    """P&Lデータを評価してレポートを生成する"""
    revenue = pl_data['revenue']
    items   = pl_data['items']
    
    if revenue <= 0:
        return {'error': '売上が取得できませんでした'}
    
    # カテゴリ別集計
    cogs     = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'cogs')
    labor    = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'labor')
    fixed    = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'fixed')
    selling  = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'selling')
    other    = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'other')
    tax      = sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'tax')
    financing= sum(v for k, v in items.items() if PL_CATEGORY.get(k) == 'financing')
    
    total_operating = cogs + labor + fixed + selling + other + tax
    total_all       = total_operating + financing
    
    gross_profit     = revenue - cogs
    operating_profit = revenue - total_operating
    net_approx       = revenue - total_all
    
    gross_margin    = gross_profit / revenue
    op_margin       = operating_profit / revenue
    cogs_ratio      = cogs / revenue
    labor_ratio     = labor / revenue
    fixed_ratio     = fixed / revenue
    
    # ─── KPI評価 ───────────────────────────────────────────────
    kpi_results = []
    
    def judge_ratio(label, actual, bench):
        higher_better = bench.get('higher_is_better', False)
        if higher_better:
            if actual >= bench['good']:   status = 'GOOD'
            elif actual >= bench['median']: status = 'OK'
            elif actual >= bench['warn']:   status = 'WARN'
            else:                           status = 'BAD'
        else:
            if actual <= bench['good']:   status = 'GOOD'
            elif actual <= bench['median']: status = 'OK'
            elif actual <= bench['warn']:   status = 'WARN'
            else:                           status = 'BAD'
        return {'label': label, 'actual': actual, 'bench': bench, 'status': status}
    
    kpi_results.append(judge_ratio('粗利率',    gross_margin, BENCHMARKS['粗利率']))
    kpi_results.append(judge_ratio('仕入比率',  cogs_ratio,   BENCHMARKS['仕入比率']))
    kpi_results.append(judge_ratio('人件費比率', labor_ratio,  BENCHMARKS['人件費比率']))
    kpi_results.append(judge_ratio('固定費比率', fixed_ratio,  BENCHMARKS['固定費比率']))
    kpi_results.append(judge_ratio('営業利益率', op_margin,    BENCHMARKS['営業利益率']))
    
    # ─── 総合スコア ───────────────────────────────────────────────
    score_map = {'GOOD': 25, 'OK': 15, 'WARN': 5, 'BAD': 0}
    score = sum(score_map[k['status']] for k in kpi_results)
    # ボーナス: 営業利益率が30%超
    if op_margin >= 0.30:
        score = min(100, score + 10)
    
    if score >= 90:   verdict = ('🏆 非常に優秀', 'excellent', '#155724')
    elif score >= 70: verdict = ('✅ 良好', 'good', '#1b5e20')
    elif score >= 50: verdict = ('⚠️ 普通', 'ok', '#7F6000')
    elif score >= 30: verdict = ('🔶 要改善', 'warn', '#E65100')
    else:             verdict = ('🚨 要対策', 'bad', '#B71C1C')
    
    # ─── 銀行データとの突合 ───────────────────────────────────────
    bank_summary = None
    if bank_records:
        period = pl_data.get('period', '')
        m = _re.match(r'(\d{4})年(\d{1,2})月', period)
        if m:
            yr, mo = int(m.group(1)), int(m.group(2))
            month_recs = [r for r in bank_records if r['year'] == yr and r['month'] == mo]
            if month_recs:
                bank_in  = sum(r['amount_in']  for r in month_recs)
                bank_out = sum(r['amount_out'] for r in month_recs)
                bank_net = bank_in - bank_out
                end_bal  = month_recs[-1]['balance']
                diff_rev = abs(revenue - bank_in)
                match_pct = max(0, 1 - diff_rev / max(revenue, 1)) * 100
                bank_summary = {
                    'bank_in': bank_in, 'bank_out': bank_out,
                    'bank_net': bank_net, 'end_balance': end_bal,
                    'diff_from_pl': diff_rev, 'match_pct': match_pct,
                    'count': len(month_recs),
                }
    
    # ─── 改善アドバイス ───────────────────────────────────────────
    advice = []
    if cogs_ratio > BENCHMARKS['仕入比率']['warn']:
        advice.append(('仕入', f'仕入比率が{cogs_ratio*100:.1f}%と高め。発注量や仕入先の見直しを検討してください。'))
    if labor_ratio > BENCHMARKS['人件費比率']['warn']:
        advice.append(('人件費', f'人件費比率が{labor_ratio*100:.1f}%。売上増加か業務効率化で吸収できるか確認を。'))
    if fixed_ratio > BENCHMARKS['固定費比率']['warn']:
        advice.append(('固定費', f'固定費比率が{fixed_ratio*100:.1f}%。家賃・通信費の見直し余地があります。'))
    if op_margin < BENCHMARKS['営業利益率']['warn']:
        advice.append(('利益', f'営業利益率が{op_margin*100:.1f}%。収益構造の根本的な見直しが必要です。'))
    if not advice:
        advice.append(('総評', '全KPIが基準値内です。この水準を維持・向上させることを目指してください。'))
    
    return {
        'period': pl_data['period'],
        'revenue': revenue,
        'cogs': cogs, 'labor': labor, 'fixed': fixed,
        'selling': selling, 'other': other, 'tax': tax, 'financing': financing,
        'gross_profit': gross_profit,
        'operating_profit': operating_profit,
        'net_approx': net_approx,
        'gross_margin': gross_margin,
        'op_margin': op_margin,
        'cogs_ratio': cogs_ratio,
        'labor_ratio': labor_ratio,
        'fixed_ratio': fixed_ratio,
        'total_operating': total_operating,
        'total_all': total_all,
        'items': items,
        'kpi_results': kpi_results,
        'score': score,
        'verdict': verdict,
        'advice': advice,
        'bank_summary': bank_summary,
    }


HTML = r'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>E-MIETA・銀行仕訳 | スマート仕訳EXCEL×MIETA経営健康診断</title>
<link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }

  @keyframes fadeIn { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: translateY(0); } }
  @keyframes pulse { 0%, 100% { transform: scale(1); } 50% { transform: scale(1.05); } }
  @keyframes sway { 0%, 100% { transform: translateX(-2px) rotate(0deg); } 50% { transform: translateX(2px) rotate(0.5deg); } }
  @keyframes float { 0%, 100% { transform: translateY(0px); } 50% { transform: translateY(-8px); } }

  body {
    font-family: 'Segoe UI', 'Hiragino Kaku Gothic ProN', 'Yu Gothic', sans-serif;
    background: white;
    min-height: 100vh; display: flex; align-items: flex-start;
    justify-content: center; padding: 30px 20px;
    position: relative;
  }

  body::before {
    content: '';
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background: transparent;
    pointer-events: none;
    z-index: -1;
  }

  .container {
    background: white;
    border-radius: 28px;
    box-shadow: 0 10px 40px rgba(0, 0, 0, 0.06);
    width: 100%;
    max-width: 720px;
    overflow: hidden;
    animation: fadeIn 0.5s ease-out;
    position: relative;
    border: 6px solid;
    border-image: linear-gradient(135deg, #00bcd4 0%, #4ade80 100%) 1;
  }

  .header {
    background: white;
    padding: 24px 28px;
    position: relative;
    overflow: hidden;
  }

  .header-top {
    display: flex;
    align-items: center;
    gap: 24px;
  }

  .character-logo {
    height: 110px;
    width: auto;
    flex-shrink: 0;
    object-fit: contain;
  }

  .title-logo {
    max-height: 90px;
    width: auto;
    display: block;
    object-fit: contain;
  }

  .header-title .subtitle {
    font-size: 13px;
    color: #666;
    margin-top: 8px;
    font-weight: 500;
  }

  .header::before {
    display: none;
  }

  .header::after {
    display: none;
  }

  .logo-container {
    display: none;
  }

  .logo-svg {
    display: none;
  }

  .character-corner {
    display: none;
  }

  /* タブ */
  .tabs {
    display: flex;
    border-bottom: 1px solid #e8e8e8;
    background: white;
    padding: 0 8px;
  }

  .tab {
    flex: 1;
    padding: 16px 8px;
    text-align: center;
    font-size: 14px;
    font-weight: 600;
    cursor: pointer;
    color: #5b9a83;
    border-bottom: 3px solid transparent;
    margin-bottom: -2px;
    transition: all 0.3s ease;
    position: relative;
  }

  .tab:hover { color: #5fa972; }

  .tab.active {
    color: #7ac97f;
    border-bottom-color: #7ac97f;
    background: white;
  }

  .tab-icon {
    width: 24px;
    height: 24px;
    display: block;
    margin: 0 auto 4px;
    transition: transform 0.3s ease;
    stroke: currentColor;
    color: inherit;
  }

  .tab.active .tab-icon { transform: scale(1.15); }

  .section-icon {
    width: 18px;
    height: 18px;
    display: inline-block;
    margin-right: 6px;
    vertical-align: -3px;
    stroke: currentColor;
    color: inherit;
  }

  .feature-icon {
    width: 20px;
    height: 20px;
    display: inline-block;
    margin-right: 8px;
    vertical-align: -3px;
    stroke: currentColor;
    color: inherit;
  }

  .upload-icon {
    width: 48px;
    height: 48px;
    display: block;
    margin: 0 auto 8px;
    stroke: #5fa972;
    color: #5fa972;
  }

  .check-icon {
    width: 20px;
    height: 20px;
    display: inline-block;
    margin-right: 6px;
    vertical-align: -2px;
    stroke: #5fa972;
    color: #5fa972;
  }

  .hint-icon {
    width: 16px;
    height: 16px;
    display: inline-block;
    margin-right: 4px;
    vertical-align: -2px;
    stroke: #ffa500;
    color: #ffa500;
  }

  .btn-icon {
    width: 18px;
    height: 18px;
    display: inline-block;
    margin-right: 6px;
    vertical-align: -3px;
    stroke: currentColor;
    color: currentColor;
  }

  .btn-icon-sm {
    width: 16px;
    height: 16px;
    display: inline-block;
    margin-right: 4px;
    vertical-align: -2px;
    stroke: currentColor;
    color: currentColor;
  }

  .panel {
    display: none;
    padding: 28px;
    animation: fadeIn 0.3s ease-out;
  }

  .panel.active { display: block; }

  /* 共通 */
  .section-title {
    font-size: 15px;
    font-weight: 700;
    color: #5fa972;
    margin-bottom: 14px;
    padding-bottom: 8px;
    border-bottom: 2px solid #e8e8e8;
    display: flex;
    align-items: center;
    gap: 8px;
  }

  .upload-area {
    border: 2.5px dashed #82c792;
    border-radius: 18px;
    padding: 32px;
    text-align: center;
    cursor: pointer;
    transition: all 0.3s ease;
    background: linear-gradient(135deg, #f0faf7 0%, #e8f5f0 100%);
    margin-bottom: 18px;
    position: relative;
  }

  .upload-area:hover {
    border-color: #7ac97f;
    background: linear-gradient(135deg, #e8f8f3 0%, #dff3eb 100%);
    transform: translateY(-2px);
  }

  .upload-area.drag {
    border-color: #7ac97f;
    background: linear-gradient(135deg, #e8f8f3 0%, #dff3eb 100%);
    box-shadow: 0 8px 20px rgba(58, 159, 93, 0.15);
  }

  .upload-area input[type=file] {
    position: absolute;
    inset: 0;
    opacity: 0;
    cursor: pointer;
  }

  .upload-icon {
    font-size: 40px;
    margin-bottom: 10px;
    animation: pulse 2s infinite;
  }

  .upload-text {
    font-size: 15px;
    color: #5fa972;
    font-weight: 600;
  }

  .upload-sub {
    font-size: 12px;
    color: #6ba88a;
    margin-top: 6px;
  }

  .file-info {
    background: linear-gradient(135deg, #d0f0d8 0%, #b8e6c8 100%);
    border-radius: 12px;
    padding: 12px 16px;
    display: none;
    align-items: center;
    gap: 10px;
    margin-bottom: 16px;
    font-size: 13px;
    color: #1f5a38;
    font-weight: 600;
  }

  .file-info.show { display: flex; }

  .btn {
    width: 100%;
    padding: 16px;
    background: linear-gradient(135deg, #7ac97f, #4fb573);
    color: white;
    border: none;
    border-radius: 14px;
    font-size: 16px;
    font-weight: 700;
    cursor: pointer;
    transition: all 0.3s ease;
    box-shadow: 0 4px 15px rgba(58, 159, 93, 0.25);
  }

  .btn:hover:not(:disabled) {
    transform: translateY(-3px);
    box-shadow: 0 8px 25px rgba(58, 159, 93, 0.35);
  }

  .btn:active:not(:disabled) { transform: translateY(-1px); }

  .btn:disabled {
    opacity: 0.5;
    cursor: not-allowed;
    transform: none;
  }

  .btn-eval {
    background: linear-gradient(135deg, #4a9fc8, #2d7a9a);
    box-shadow: 0 4px 15px rgba(30, 90, 122, 0.25);
  }

  .btn-eval:hover:not(:disabled) {
    box-shadow: 0 8px 25px rgba(30, 90, 122, 0.35);
  }

  .progress {
    display: none;
    margin-top: 18px;
    background: #d0e8e0;
    border-radius: 100px;
    height: 8px;
    overflow: hidden;
  }

  .progress.show { display: block; }

  .progress-bar {
    height: 100%;
    width: 0%;
    background: linear-gradient(90deg, #7ac97f, #4fb573, #4a9fc8);
    border-radius: 100px;
    transition: width 0.3s ease;
    box-shadow: 0 0 10px rgba(58, 159, 93, 0.4);
  }

  .status {
    text-align: center;
    font-size: 13px;
    color: #5b9a83;
    margin-top: 12px;
    min-height: 20px;
    font-weight: 500;
  }

  .result {
    display: none;
    margin-top: 20px;
    background: linear-gradient(135deg, #d0f0d8, #b8e6c8);
    border-radius: 16px;
    padding: 24px;
    text-align: center;
  }

  .result.show { display: block; }

  .dl-btn {
    display: inline-block;
    padding: 14px 32px;
    background: linear-gradient(135deg, #7ac97f, #4fb573);
    color: white;
    border-radius: 12px;
    text-decoration: none;
    font-weight: 700;
    font-size: 15px;
    transition: all 0.3s ease;
    box-shadow: 0 4px 15px rgba(58, 159, 93, 0.25);
  }

  .dl-btn:hover {
    background: linear-gradient(135deg, #2d8850, #3fa065);
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(58, 159, 93, 0.35);
  }

  .error-msg {
    display: none;
    margin-top: 14px;
    padding: 14px;
    border-radius: 12px;
    background: linear-gradient(135deg, #ffebee, #ffcdd2);
    color: #c62828;
    font-size: 13px;
    font-weight: 500;
  }

  .error-msg.show { display: block; }

  /* P&L テキスト入力 */
  .pl-textarea {
    width: 100%;
    height: 260px;
    padding: 14px;
    border: 2px solid #d8d8d8;
    border-radius: 14px;
    font-size: 13px;
    font-family: 'Courier New', monospace;
    resize: vertical;
    transition: all 0.3s ease;
    background: white;
  }

  .pl-textarea:focus {
    outline: none;
    border-color: #7ac97f;
    box-shadow: 0 0 0 3px rgba(58, 159, 93, 0.1);
  }

  .hint {
    font-size: 12px;
    color: #6ba88a;
    margin-top: 8px;
    margin-bottom: 16px;
    font-weight: 500;
  }

  /* 評価結果 */
  .eval-result { display: none; }
  .eval-result.show { display: block; margin-top: 24px; }

  .verdict-box {
    border-radius: 18px;
    padding: 24px;
    margin-bottom: 22px;
    text-align: center;
    box-shadow: 0 4px 15px rgba(0,0,0,0.08);
  }

  .verdict-excellent { background: linear-gradient(135deg, #d0f0d8, #9dd9b8); border: 2px solid #5fa972; }
  .verdict-good      { background: linear-gradient(135deg, #d0f0d8, #b8e6c8); border: 2px solid #388e3c; }
  .verdict-ok        { background: linear-gradient(135deg, #fff8d8, #fff0a8); border: 2px solid #f9a825; }
  .verdict-warn      { background: linear-gradient(135deg, #ffe8c8, #ffd9a8); border: 2px solid #ef6c00; }
  .verdict-bad       { background: linear-gradient(135deg, #ffebee, #ffcdd2); border: 2px solid #c62828; }

  .verdict-label {
    font-size: 32px;
    font-weight: 800;
    margin-bottom: 6px;
  }

  .verdict-score {
    font-size: 15px;
    opacity: 0.9;
    font-weight: 600;
  }

  .verdict-period {
    font-size: 14px;
    font-weight: 600;
    margin-bottom: 10px;
    opacity: 0.8;
  }

  /* KPI表 */
  .kpi-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
    margin-bottom: 22px;
  }

  .kpi-card {
    border-radius: 14px;
    padding: 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    transition: all 0.3s ease;
  }

  .kpi-card:hover { transform: translateY(-2px); }

  .kpi-card.GOOD { background: linear-gradient(135deg, #d0f0d8, #b8e6c8); border-left: 5px solid #5fa972; }
  .kpi-card.OK   { background: linear-gradient(135deg, #fff8d8, #fff0a8); border-left: 5px solid #f9a825; }
  .kpi-card.WARN { background: linear-gradient(135deg, #ffe8c8, #ffd9a8); border-left: 5px solid #ef6c00; }
  .kpi-card.BAD  { background: linear-gradient(135deg, #ffebee, #ffcdd2); border-left: 5px solid #c62828; }

  .kpi-name  {
    font-size: 12px;
    color: #556;
    margin-bottom: 6px;
    font-weight: 500;
  }

  .kpi-value {
    font-size: 24px;
    font-weight: 800;
  }

  .kpi-value.GOOD { color: #5fa972; }
  .kpi-value.OK   { color: #f9a825; }
  .kpi-value.WARN { color: #ef6c00; }
  .kpi-value.BAD  { color: #c62828; }

  .kpi-bench {
    font-size: 11px;
    color: #888;
    margin-top: 4px;
    font-weight: 500;
  }

  .kpi-badge {
    font-size: 11px;
    font-weight: 700;
    padding: 3px 8px;
    border-radius: 100px;
    float: right;
  }

  .kpi-badge.GOOD { background: #5fa972; color: white; }
  .kpi-badge.OK   { background: #f9a825; color: white; }
  .kpi-badge.WARN { background: #ef6c00; color: white; }
  .kpi-badge.BAD  { background: #c62828; color: white; }

  /* P&L表 */
  .pl-table {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 22px;
    font-size: 13px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    border-radius: 10px;
    overflow: hidden;
  }

  .pl-table th {
    background: linear-gradient(135deg, #5fa972, #7ac97f);
    color: white;
    padding: 10px 12px;
    text-align: left;
    font-weight: 600;
  }

  .pl-table td {
    padding: 8px 12px;
    border-bottom: 1px solid #e8e8e8;
  }

  .pl-table tr:last-child td { border-bottom: none; }

  .pl-table .subtotal td {
    background: #f5f5f5;
    font-weight: 600;
  }

  .pl-table .profit td {
    background: linear-gradient(135deg, #d0f0d8, #b8e6c8);
    font-weight: 700;
    color: #1f5a38;
  }

  .pl-table .profit-neg td {
    background: linear-gradient(135deg, #ffebee, #ffe8ec);
    font-weight: 700;
    color: #c62828;
  }

  .pl-table .right {
    text-align: right;
    font-variant-numeric: tabular-nums;
  }

  .pl-table .ratio {
    color: #888;
    font-size: 11px;
  }

  /* 銀行突合 */
  .bank-match {
    background: white;
    border: 1px solid #e0e0e0;
    border-radius: 14px;
    padding: 18px;
    margin-bottom: 22px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
  }

  .bank-match h4 {
    font-size: 14px;
    font-weight: 700;
    color: #5fa972;
    margin-bottom: 12px;
  }

  .match-row {
    display: flex;
    justify-content: space-between;
    font-size: 13px;
    padding: 6px 0;
    border-bottom: 1px solid #e8e8e8;
  }

  .match-row:last-child { border-bottom: none; }

  /* アドバイス */
  .advice-box {
    border-radius: 14px;
    padding: 16px;
    background: white;
    margin-bottom: 10px;
    border-left: 4px solid #7ac97f;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    transition: all 0.3s ease;
  }

  .advice-box:hover { transform: translateX(2px); }

  .advice-title {
    font-size: 13px;
    font-weight: 700;
    color: #5fa972;
    margin-bottom: 6px;
  }

  .advice-text {
    font-size: 13px;
    color: #555;
    line-height: 1.7;
    font-weight: 500;
  }

  /* features */
  .features {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px;
    margin-bottom: 22px;
  }

  .feature {
    font-size: 13px;
    color: #5fa972;
    display: flex;
    align-items: center;
    gap: 8px;
    background: #f8f8f8;
    padding: 10px 14px;
    border-radius: 12px;
    font-weight: 500;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    transition: all 0.3s ease;
  }

  .feature:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
  }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <div class="header-top">
      <img src="/static/samo-char.png" alt="SAMO" class="character-logo">
      <div class="header-title">
        <img src="/static/emieta-logo.png" alt="E-MIETA" class="title-logo">
        <p class="subtitle">スマート仕訳×MIETA経営健康診断</p>
      </div>
    </div>
  </div>

  <div class="tabs">
    <div class="tab active" onclick="switchTab('excel')">
      <svg class="tab-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="2" x2="12" y2="22"></line><path d="M17 5H9.5a1.5 1.5 0 0 0-1.5 1.5v12a1.5 1.5 0 0 0 1.5 1.5H17"></path><path d="M6 12H2m20 0h-4"></path></svg><span>CSV変換</span>
    </div>
    <div class="tab" onclick="switchTab('eval')">
      <svg class="tab-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"></circle><path d="m21 21-4.35-4.35"></path></svg><span>経営分析</span>
    </div>
  </div>

  <!-- ===== TAB1: Excel変換 ===== -->
  <div class="panel active" id="panel-excel">
    <div class="features">
      <div class="feature"><svg class="feature-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="4" width="18" height="18" rx="2" ry="2"></rect><line x1="16" y1="2" x2="16" y2="6"></line><line x1="8" y1="2" x2="8" y2="6"></line><line x1="3" y1="10" x2="21" y2="10"></line></svg> 月別自動振分け</div>
      <div class="feature"><svg class="feature-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20.59 13.41l-7.17 7.17a2 2 0 0 1-2.83 0L2 12V2h10l8.59 8.59a2 2 0 0 1 0 2.82z"></path><line x1="7" y1="7" x2="7.01" y2="7"></line></svg> 科目自動判定</div>
      <div class="feature"><svg class="feature-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"></circle><path d="M12 7v10m-3-3h6"></path></svg> 数式計算対応</div>
      <div class="feature"><svg class="feature-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="2" x2="12" y2="22"></line><path d="M17 5H9.5a1.5 1.5 0 0 0-1.5 1.5v12a1.5 1.5 0 0 0 1.5 1.5H17"></path><path d="M6 12H2m20 0h-4"></path></svg> 集計シート生成</div>
    </div>

    <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline></svg> 銀行明細CSVをアップロード</div>
    <div class="upload-area" id="dropZone">
      <input type="file" id="fileInput" accept=".csv">
      <svg class="upload-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="17 8 12 3 7 8"></polyline><line x1="12" y1="3" x2="12" y2="15"></line></svg>
      <div class="upload-text">CSVファイルをここにドロップ</div>
      <div class="upload-sub">またはクリックして選択（Shift-JIS/UTF-8 自動判定）</div>
    </div>
    <div class="file-info" id="fileInfo"><svg class="check-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg><span id="fileName"></span></div>
    <button class="btn" id="convertBtn" disabled onclick="convert()">✈️ Excelを生成する</button>
    <div class="progress" id="progress"><div class="progress-bar" id="progressBar"></div></div>
    <div class="status" id="status"></div>
    <div class="error-msg" id="errorMsg"></div>
    <div class="result" id="result">
      <div style="font-size:48px;margin-bottom:12px">✅</div>
      <div style="font-weight:700;font-size:17px;color:#1b5e20;margin-bottom:4px">Excel生成完了！</div>
      <div style="font-size:12px;color:#666;margin-bottom:16px">すぐにダウンロードできます</div>
      <a id="dlLink" class="dl-btn" href="#" download"><svg class="btn-icon-sm" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg> Excelをダウンロード</a>
    </div>

    <!-- 使い方ガイド -->
    <div style="margin-top:28px; padding-top:28px; border-top: 2px solid #f0e6f8;">
      <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg> CSV変換の使い方</div>
      <div class="advice-box">
        <div class="advice-title">【ステップ1】対応するCSV形式を確認</div>
        <div class="advice-text">
          銀行明細は以下の列構成に対応しています：<br>
          <strong>日付 | 摘要 | 入金金額 | 出金金額 | 残高 | メモ</strong><br>
          日付は「YYYYMMDD」「YYYY/MM/DD」「YYYY-MM-DD」形式に対応
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【ステップ2】ファイルをアップロード</div>
        <div class="advice-text">
          銀行明細CSVファイルをアップロードエリアにドロップするか、クリックして選択してください。Shift-JISとUTF-8は自動判定されます。
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【ステップ3】Excel生成ボタンをクリック</div>
        <div class="advice-text">
          「Excelを生成する」ボタンをクリックすると、以下が自動生成されます：<br>
          • 年間サマリーシート（月別集計）<br>
          • 各月シート（月別詳細明細）<br>
          • カテゴリ別集計シート
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【ステップ4】Excelファイルをダウンロード</div>
        <div class="advice-text">
          生成が完了したら「Excelをダウンロード」ボタンからファイルをダウンロードできます。Excel内の数式は自動更新されます。
        </div>
      </div>
    </div>
  </div>

  <!-- ===== TAB2: 経営分析 ===== -->
  <div class="panel" id="panel-eval">
    <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg> 月次P&L内訳を貼り付け</div>
    <textarea class="pl-textarea" id="plText" placeholder="例：
2026年 2月
売上　13,416,660

仕入　　　　3,485,000
外注費　　　125,000
人件費　　　1,389,000
交際費　　　17,000
旅費交通費　250,000
（以下つづき...）"></textarea>
    <div class="hint"><svg class="hint-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg> 会計ソフトなどからP&Lをコピペするだけで分析できます</div>

    <div class="section-title" style="margin-top:20px"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"></circle><path d="M12 7v10m-3-3h6"></path></svg> 銀行明細CSV（オプション）</div>
    <div class="upload-area" id="dropZone2">
      <input type="file" id="fileInput2" accept=".csv">
      <svg class="upload-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="17 8 12 3 7 8"></polyline><line x1="12" y1="3" x2="12" y2="15"></line></svg>
      <div class="upload-text">CSVをドロップ（省略可）</div>
      <div class="upload-sub">入力すると銀行残高と自動照合します</div>
    </div>
    <div class="file-info" id="fileInfo2"><svg class="check-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg><span id="fileName2"></span></div>

    <button class="btn btn-eval" id="evalBtn" onclick="evaluate()"><svg class="btn-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"></circle><path d="m21 21-4.35-4.35"></path></svg> 経営分析を実行</button>
    <div class="progress" id="progress2"><div class="progress-bar" id="progressBar2"></div></div>
    <div class="status" id="status2"></div>
    <div class="error-msg" id="errorMsg2"></div>

    <!-- 評価結果 -->
    <div class="eval-result" id="evalResult"></div>

    <!-- 使い方ガイド -->
    <div style="margin-top:28px; padding-top:28px; border-top: 2px solid #f0e6f8;">
      <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg> 経営分析の使い方</div>
      <div class="advice-box">
        <div class="advice-title">【分析結果の見方】総合スコア</div>
        <div class="advice-text">
          <strong>◎ 優（85～100点）</strong>：経営状況が優秀。このペースを維持してください。<br>
          <strong>○ 良（70～84点）</strong>：経営状況が良好。さらなる改善の余地あり。<br>
          <strong>△ 注（55～69点）</strong>：注意が必要。改善策を実行しましょう。<br>
          <strong>✕ 要（0～54点）</strong>：即座の改善が必要です。アドバイスを参考に。
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【KPI指標の意味】</div>
        <div class="advice-text">
          <strong>粗利率</strong>：売上から原価を引いた利益の割合。高いほど良好。<br>
          <strong>営業利益率</strong>：売上から全ての費用を引いた利益の割合。企業の稼ぐ力を示します。<br>
          <strong>人件費率</strong>：売上に占める人件費の割合。適切な範囲内が理想的。<br>
          <strong>固定費率</strong>：売上に占める固定費の割合。低いほど経営の効率性が高い。
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【損益計算書の読み方】</div>
        <div class="advice-text">
          売上から順に原価、経費が差し引かれて、最終的な利益が表示されます。<br>
          「売上比」の列は、各項目が売上に占める割合。<br>
          <span style="background:#d4f0da;padding:2px 6px;border-radius:4px;display:inline-block;margin-top:6px;font-weight:600;color:#1b5e20;">green=利益</span>
          <span style="background:#ffe8ec;padding:2px 6px;border-radius:4px;display:inline-block;margin-left:6px;font-weight:600;color:#c62828;">red=損失</span>
        </div>
      </div>
      <div class="advice-box">
        <div class="advice-title">【銀行データとの照合】</div>
        <div class="advice-text">
          CSV銀行明細をアップロードすると、P&L売上と銀行の入金額を比較できます。<br>
          差額が出た場合は、未計上や入力ミスの可能性があります。
        </div>
      </div>
    </div>
  </div>
</div>

<script>
// ===== タブ切り替え =====
function switchTab(name) {
  document.querySelectorAll('.tab').forEach((t,i) => t.classList.toggle('active', ['excel','eval'][i] === name));
  document.querySelectorAll('.panel').forEach((p,i) => p.classList.toggle('active', ['panel-excel','panel-eval'][i] === 'panel-'+name));
}

// ===== Tab1: Excel変換 =====
const fileInput  = document.getElementById('fileInput');
const dropZone   = document.getElementById('dropZone');
const fileInfo   = document.getElementById('fileInfo');
const fileName   = document.getElementById('fileName');
const convertBtn = document.getElementById('convertBtn');
let selectedFile = null;

fileInput.addEventListener('change', e => handleFile(e.target.files[0]));
dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('drag'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('drag'));
dropZone.addEventListener('drop', e => { e.preventDefault(); dropZone.classList.remove('drag'); handleFile(e.dataTransfer.files[0]); });

function handleFile(file) {
  if (!file) return;
  selectedFile = file;
  fileName.textContent = `${file.name}（${(file.size/1024).toFixed(0)} KB）`;
  fileInfo.classList.add('show');
  convertBtn.disabled = false;
  document.getElementById('result').classList.remove('show');
  document.getElementById('errorMsg').classList.remove('show');
}

async function convert() {
  if (!selectedFile) return;
  convertBtn.disabled = true;
  const progress = document.getElementById('progress');
  const bar = document.getElementById('progressBar');
  progress.classList.add('show');
  let pct = 0;
  const timer = setInterval(() => {
    pct = Math.min(pct + Math.random()*12, 88);
    bar.style.width = pct + '%';
    const msgs = ['📥 CSV読み込み中...','📊 月別整理中...','📋 シート生成中...','🏥 診断中...'];
    document.getElementById('status').textContent = msgs[Math.floor(pct/25)] || msgs[3];
  }, 200);
  const fd = new FormData(); fd.append('file', selectedFile);
  try {
    const res = await fetch('/convert', {method:'POST', body:fd});
    clearInterval(timer); bar.style.width = '100%';
    if (!res.ok) { const e = await res.json(); throw new Error(e.error); }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    document.getElementById('dlLink').href = url;
    const cd = res.headers.get('Content-Disposition') || '';
    const m = cd.match(/filename\*=UTF-8''(.+)/);
    document.getElementById('dlLink').download = m ? decodeURIComponent(m[1]) : 'E-MEITA仕訳Excel.xlsx';
    document.getElementById('status').textContent = '✅ 完了！';
    document.getElementById('result').classList.add('show');
  } catch(e) {
    clearInterval(timer); bar.style.width = '0%';
    document.getElementById('errorMsg').textContent = '❌ ' + e.message;
    document.getElementById('errorMsg').classList.add('show');
    document.getElementById('status').textContent = '';
  }
  setTimeout(() => { progress.classList.remove('show'); }, 1500);
  convertBtn.disabled = false;
}

// ===== Tab2: 経営評価 =====
const fileInput2  = document.getElementById('fileInput2');
const dropZone2   = document.getElementById('dropZone2');
const fileInfo2   = document.getElementById('fileInfo2');
const fileName2   = document.getElementById('fileName2');
let csvFile2 = null;

fileInput2.addEventListener('change', e => handleFile2(e.target.files[0]));
dropZone2.addEventListener('dragover', e => { e.preventDefault(); dropZone2.classList.add('drag'); });
dropZone2.addEventListener('dragleave', () => dropZone2.classList.remove('drag'));
dropZone2.addEventListener('drop', e => { e.preventDefault(); dropZone2.classList.remove('drag'); handleFile2(e.dataTransfer.files[0]); });

function handleFile2(file) {
  if (!file) return;
  csvFile2 = file;
  fileName2.textContent = `${file.name}（${(file.size/1024).toFixed(0)} KB）`;
  fileInfo2.classList.add('show');
}

function fmt(n) { return n >= 0 ? '¥' + n.toLocaleString() : '▲¥' + Math.abs(n).toLocaleString(); }
function pct(v) { return (v*100).toFixed(1) + '%'; }
function statusLabel(s) { return {GOOD:'◎ 良好', OK:'○ 普通', WARN:'△ 注意', BAD:'✕ 要改善'}[s]; }

async function evaluate() {
  const text = document.getElementById('plText').value.trim();
  if (!text) { alert('P&L内訳テキストを貼り付けてください'); return; }
  
  document.getElementById('evalBtn').disabled = true;
  document.getElementById('progress2').classList.add('show');
  document.getElementById('evalResult').classList.remove('show');
  document.getElementById('errorMsg2').classList.remove('show');
  let pct2 = 0;
  const timer = setInterval(() => {
    pct2 = Math.min(pct2 + 15, 88);
    document.getElementById('progressBar2').style.width = pct2 + '%';
    document.getElementById('status2').textContent = pct2 < 40 ? '📊 数値解析中...' : pct2 < 70 ? '🔍 ベンチマーク比較中...' : '🏥 診断レポート生成中...';
  }, 250);

  const fd = new FormData();
  fd.append('pl_text', text);
  if (csvFile2) fd.append('csv_file', csvFile2);

  try {
    const res = await fetch('/evaluate', {method:'POST', body:fd});
    clearInterval(timer);
    document.getElementById('progressBar2').style.width = '100%';
    if (!res.ok) { const e = await res.json(); throw new Error(e.error); }
    const data = await res.json();
    renderEvalResult(data);
    document.getElementById('status2').textContent = '✅ 評価完了！';
  } catch(e) {
    clearInterval(timer);
    document.getElementById('errorMsg2').textContent = '❌ ' + e.message;
    document.getElementById('errorMsg2').classList.add('show');
    document.getElementById('status2').textContent = '';
  }
  setTimeout(() => document.getElementById('progress2').classList.remove('show'), 1000);
  document.getElementById('evalBtn').disabled = false;
}

function renderEvalResult(d) {
  const verdictClass = {excellent:'verdict-excellent', good:'verdict-good', ok:'verdict-ok', warn:'verdict-warn', bad:'verdict-bad'}[d.verdict[1]];
  const verdictColor = d.verdict[2];
  
  let html = `
  <div class="verdict-box ${verdictClass}">
    <div class="verdict-period">${d.period}</div>
    <div class="verdict-label" style="color:${verdictColor}">${d.verdict[0]}</div>
    <div class="verdict-score">総合スコア ${d.score} / 100点</div>
  </div>

  <div class="kpi-grid">`;

  for (const k of d.kpi_results) {
    const bench = k.bench;
    const isHB = bench.higher_is_better;
    const bLabel = isHB ? `目安 ${(bench.median*100).toFixed(0)}%以上` : `目安 ${(bench.median*100).toFixed(0)}%以下`;
    html += `
    <div class="kpi-card ${k.status}">
      <div class="kpi-name">${bench.label} <span class="kpi-badge ${k.status}">${statusLabel(k.status)}</span></div>
      <div class="kpi-value ${k.status}">${(k.actual*100).toFixed(1)}%</div>
      <div class="kpi-bench">${bLabel}</div>
    </div>`;
  }
  html += `</div>`;

  // P&L表
  html += `
  <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"></path><rect x="8" y="2" width="8" height="4" rx="1" ry="1"></rect></svg> 損益計算書（簡易）</div>
  <table class="pl-table">
    <tr><th>項目</th><th class="right">金額</th><th class="right">売上比</th></tr>
    <tr><td>売上</td><td class="right">${fmt(d.revenue)}</td><td class="right ratio">100.0%</td></tr>
    <tr><td>　仕入・原価</td><td class="right">${fmt(d.cogs)}</td><td class="right ratio">${pct(d.cogs_ratio)}</td></tr>
    <tr class="profit"><td>粗利益</td><td class="right">${fmt(d.gross_profit)}</td><td class="right">${pct(d.gross_margin)}</td></tr>
    <tr><td>　人件費・外注費</td><td class="right">${fmt(d.labor)}</td><td class="right ratio">${pct(d.labor/d.revenue)}</td></tr>
    <tr><td>　固定費</td><td class="right">${fmt(d.fixed)}</td><td class="right ratio">${pct(d.fixed_ratio)}</td></tr>
    <tr><td>　その他経費</td><td class="right">${fmt(d.selling+d.other+d.tax)}</td><td class="right ratio">${pct((d.selling+d.other+d.tax)/d.revenue)}</td></tr>
    <tr class="${d.operating_profit >= 0 ? 'profit' : 'profit-neg'}"><td>営業利益</td><td class="right">${fmt(d.operating_profit)}</td><td class="right">${pct(d.op_margin)}</td></tr>
    <tr><td>　財務費用（借入等）</td><td class="right">${fmt(d.financing)}</td><td class="right ratio">${pct(d.financing/d.revenue)}</td></tr>
    <tr class="${d.net_approx >= 0 ? 'profit' : 'profit-neg'}"><td>税引前利益（概算）</td><td class="right">${fmt(d.net_approx)}</td><td class="right">${pct(d.net_approx/d.revenue)}</td></tr>
  </table>`;

  // 銀行突合
  if (d.bank_summary) {
    const b = d.bank_summary;
    html += `
  <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"></circle><path d="M12 7v10m-3-3h6"></path></svg> 銀行データとの照合</div>
  <div class="bank-match">
    <h4>対象月: ${d.period}　（${b.count}件）</h4>
    <div class="match-row"><span>銀行入金合計</span><span style="font-weight:600">${fmt(b.bank_in)}</span></div>
    <div class="match-row"><span>銀行出金合計</span><span style="font-weight:600">${fmt(b.bank_out)}</span></div>
    <div class="match-row"><span>銀行ネット増減</span><span style="font-weight:600;color:${b.bank_net>=0?'#2e7d32':'#c62828'}">${fmt(b.bank_net)}</span></div>
    <div class="match-row"><span>月末残高</span><span style="font-weight:700">${fmt(b.end_balance)}</span></div>
    <div class="match-row"><span>P&L売上との差額</span><span>${fmt(b.diff_from_pl)}</span></div>
  </div>`;
  }

  // 費用明細
  html += `
  <div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h5l2 3h9a2 2 0 0 1 2 2z"></path></svg> 費用内訳（全科目）</div>
  <table class="pl-table">
    <tr><th>科目</th><th class="right">金額</th><th class="right">売上比</th></tr>`;
  for (const [k, v] of Object.entries(d.items).sort((a,b)=>b[1]-a[1])) {
    html += `<tr><td>　${k}</td><td class="right">${fmt(v)}</td><td class="right ratio">${pct(v/d.revenue)}</td></tr>`;
  }
  html += `</table>`;

  // アドバイス
  html += `<div class="section-title"><svg class="section-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path></svg> 改善アドバイス</div>`;
  for (const [title, text] of d.advice) {
    html += `<div class="advice-box"><div class="advice-title">【${title}】</div><div class="advice-text">${text}</div></div>`;
  }

  const el = document.getElementById('evalResult');
  el.innerHTML = html;
  el.classList.add('show');
  el.scrollIntoView({behavior:'smooth', block:'start'});
}
</script>
</body>
</html>'''

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/convert', methods=['POST'])
def convert():
    from flask import Response
    import urllib.parse
    
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルが見つかりません'}), 400
    
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    
    try:
        file_bytes = f.read()
        records = parse_bank_csv(file_bytes)
        
        if not records:
            return jsonify({'error': 'データが読み込めませんでした。CSVの形式を確認してください'}), 400
        
        wb = build_excel(records)
        
        # ファイル名生成
        years  = sorted(set(r['year'] for r in records))
        months = sorted(set(r['month'] for r in records))
        year_str = f"{years[0]}" if len(years) == 1 else f"{years[0]}-{years[-1]}"
        filename = f"E-MEITA仕訳Excel_{year_str}年_月別.xlsx"
        
        # メモリに保存して返す
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        encoded_name = urllib.parse.quote(filename)
        
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_name}",
                'X-Record-Count': str(len(records)),
            }
        )
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/evaluate', methods=['POST'])
def evaluate():
    pl_text = request.form.get('pl_text', '').strip()
    if not pl_text:
        return jsonify({'error': 'P&Lテキストが入力されていません'}), 400
    
    try:
        pl_data = parse_pl_text(pl_text)
        
        if pl_data['revenue'] <= 0:
            return jsonify({'error': '売上金額が取得できませんでした。「売上　13,416,660」のような形式で入力してください'}), 400
        
        # 銀行CSVが添付されていれば解析
        bank_records = None
        if 'csv_file' in request.files:
            f = request.files['csv_file']
            if f.filename:
                try:
                    bank_records = parse_bank_csv(f.read())
                except:
                    pass
        
        result = evaluate_pl(pl_data, bank_records)
        return jsonify(result)
    
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    print(f"🏦 銀行明細変換システム起動中... http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=False)
